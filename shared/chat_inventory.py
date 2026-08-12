"""Chat/mail inventory indexer — vendor-wise supply intelligence.

Scans the inbound WhatsApp store (`shared/channels`) and recent Gmail
(`shared/mailbox`), then builds a **vendor-wise** index of what each chat or
sender offers: **products / raw materials**, **equipment**, and **services** —
with the **rates** quoted, quantities, dates, evidence snippets, and deep
links back to the source mail/chat.

Outputs:
- ``data/state/chat_inventory.json`` — regenerated on every comms digest and
  on demand (``POST /api/inventory/run``), rendered by the Desktop Hub.
- ``GET /api/inventory/export.xlsx`` / the ``inventory.export`` chat tool —
  a three-sheet (Products / Equipment / Services) Excel workbook.

Extraction is heuristic-first (domain keyword sets + price patterns) so it
works with the engine busy or offline; snippets are always kept so a human
can verify any row against the source message.
"""
from __future__ import annotations

import json
import re
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from .config import ROOT_DIR

INVENTORY_PATH = ROOT_DIR / "data" / "state" / "chat_inventory.json"
EXPORT_DIR = ROOT_DIR / "data" / "media" / "exports"

# ---------------------------------------------------------------------------
# Domain keyword sets — pharma/chem trading + general manufacturing.
# ---------------------------------------------------------------------------
_MATERIAL_MARKERS = re.compile(
    r"\b(quinoline|hydroxyquinoline|api|intermediate|excipient|solvent|reagent|"
    r"powder|granules?|crystals?|flake|resin|polymer|plywood|ply|laminate|"
    r"mdf|particle\s?board|veneer|film|foil|sheet|coil|wire|pipe|valve|"
    r"chemicals?|acid|base|salt|catalyst|pigment|dye|filler|additive|"
    r"starch|cellulose|gelatin|lactose|titanium|zinc|copper|aluminium|"
    r"aluminum|steel|cas\s?no|cas\s?#)\b", re.I)
_EQUIPMENT_MARKERS = re.compile(
    r"\b(machin(e|ery)|reactor|dryer|blender|mixer|granulator|mill|sieve|"
    r"centrifuge|autoclave|boiler|chiller|compressor|pump|motor|conveyor|"
    r"filling\s?(machine|line)|sealing|packaging\s?(machine|line)|capping|"
    r"label(l)?ing|tablet\s?press|capsule|coating\s?pan|hplc|gc|uv\s?spectro|"
    r"spectrophotometer|balance|weighing|ph\s?meter|incubator|oven|furnace|"
    r"cutting\s?(machine|video)|cnc|lathe|welder|generator|forklift)\b", re.I)
_SERVICE_MARKERS = re.compile(
    r"\b(transport|logistics|courier|shipping|freight|dispatch\s?service|"
    r"calibration|validation\s?service|amc|maintenance\s?(contract|service)|"
    r"repair|installation|commissioning|job\s?work|cmo|contract\s?(manufacturing|research)|"
    r"third[- ]party\s?(lab|testing|laboratory)|testing\s?service|consulting|"
    r"consultancy|audit\s?service|custom\s?clearance|import\s?service)\b", re.I)

_QTY_RE = re.compile(
    r"\b\d+(?:\.\d+)?\s?(?:kg|g|gm|gram|mg|ml|l|ltr|litre|liter|ton|mt|pcs|"
    r"units?|boxes|bags?|drums?|mm|cm|inch|ft|sq\.?\s?ft)\b", re.I)
_OFFER_RE = re.compile(
    r"\b(ready\s?stock|available|price|rate|quote|quotation|offer|deal|"
    r"supply|supplier|manufacturer|wholesale|bulk|moq|immediate\s?dispatch|"
    r"export\s?quality|best\s?price|discount)\b", re.I)
_NOISE_SENDERS = re.compile(r"status@broadcast", re.I)

# Rates: ₹450, Rs.1200, INR 95/kg, $12, USD 3.5/kg, 2.5 lakh, 1 crore, @ 95/-
_PRICE_RE = re.compile(
    r"(?:₹|rs\.?|inr|usd|\$)\s?\d[\d,]*(?:\.\d+)?\s?(?:lakh|lac|crore|thousand|k)?(?:\s?(?:per|/)\s?\w+)?"
    r"|\d[\d,]*(?:\.\d+)?\s?(?:lakh|lac|crore|thousand|k)\b"
    r"|@\s?\d[\d,]*(?:\.\d+)?(?:\s?/-)?"
    r"|\d[\d,]*(?:\.\d+)?\s?(?:/-|per\s?(?:kg|g|gm|mg|ml|l|ltr|litre|liter|ton|mt|pc|pcs|unit|box|bag|drum))\b",
    re.I)


def _clean_rate(p: str) -> str:
    return re.sub(r"\s+", " ", p.strip().rstrip(","))


def _now_iso() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M")


def _env_int(name: str, default: int, *, lo: int = 0, hi: int = 500) -> int:
    try:
        return max(lo, min(int(os.getenv(name, str(default))), hi))
    except Exception:
        return default


def _sentences(text: str) -> list[str]:
    return [s.strip() for s in re.split(r"(?<=[.!?\n])\s+", text or "") if s.strip()]


def _category_hits(sent: str) -> list[str]:
    cats = []
    if _MATERIAL_MARKERS.search(sent):
        cats.append("products")
    if _EQUIPMENT_MARKERS.search(sent):
        cats.append("equipment")
    if _SERVICE_MARKERS.search(sent):
        cats.append("services")
    return cats


def _extract_items(text: str, *, source: str, url: str, date: str) -> list[dict[str, Any]]:
    """Item rows out of one message: category, evidence, optional rate/qty."""
    rows: list[dict[str, Any]] = []
    prices = _PRICE_RE.findall(text or "")
    for sent in _sentences(text):
        if len(sent) > 600:
            continue
        cats = _category_hits(sent)
        if not cats:
            continue
        sent_prices = _PRICE_RE.findall(sent) or prices  # sentence first, message as fallback
        qty = _QTY_RE.search(sent)
        finalized = bool(_FINALIZED_RE.search(sent) or _FINALIZED_RE.search(text[:2000]))
        for cat in cats:
            rows.append({
                "category": cat,
                "item": sent[:160],
                "rate": ", ".join(dict.fromkeys(_clean_rate(p) for p in sent_prices))[:120] if sent_prices else "",
                "qty": qty.group(0) if qty else "",
                "date": date,
                "source": source,
                "url": url,
                "offer": bool(_OFFER_RE.search(sent)),
                "finalized": finalized,
            })
    return rows


def _gather_whatsapp(limit: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[str]]]:
    from . import channels
    feed = channels.recent("whatsapp", limit)
    vendors: dict[str, dict[str, Any]] = {}
    quotes: list[dict[str, Any]] = []
    raw_texts: dict[str, list[str]] = {}
    for m in feed.get("messages") or []:
        if _NOISE_SENDERS.search(m.get("sender_id") or ""):
            continue
        is_mine = bool((m.get("metadata") or {}).get("is_mine"))
        text = m.get("text") or ""
        ts = m.get("received_at") or ""
        date = (datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d")
                if str(ts).replace(".", "").isdigit() else str(ts)[:10])
        thread = (m.get("thread_id") or "").replace("@s.whatsapp.net", "").replace("@g.us", "")
        if is_mine:
            # Outbound: the counterparty is who WE quoted.
            counter = thread or (m.get("sender_id") or "")
            url = f"https://wa.me/{counter}" if counter.isdigit() and not m.get("is_group") else ""
            for row in _extract_items(text, source="whatsapp", url=url, date=date):
                if row["rate"]:
                    row["counterparty"] = counter or "unknown"
                    quotes.append(row)
            raw_texts.setdefault(f"You → {counter or 'chat'}", []).append(text[:400])
            continue
        vendor = (m.get("sender_name") or m.get("sender_id") or "unknown").strip()
        sender_id = (m.get("sender_id") or "").strip()
        url = (f"https://wa.me/{sender_id}"
               if sender_id.isdigit() and not m.get("is_group") else "")
        entry = vendors.setdefault(vendor, {
            "vendor": vendor, "channels": {"whatsapp"}, "last_at": "",
            "products": [], "equipment": [], "services": [],
            "message_count": 0, "url": url})
        entry["message_count"] += 1
        if date > entry["last_at"]:
            entry["last_at"] = date
        if url and not entry["url"]:
            entry["url"] = url
        raw_texts.setdefault(vendor, []).append(text[:400])
        for row in _extract_items(text, source="whatsapp", url=url, date=date):
            key = row["category"]
            if row["item"] not in [r["item"] for r in entry[key]]:
                entry[key].append(row)
    return list(vendors.values()), quotes, raw_texts


def _gather_gmail(max_mail: int) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, list[str]]]:
    vendors: dict[str, dict[str, Any]] = {}
    quotes: list[dict[str, Any]] = []
    raw_texts: dict[str, list[str]] = {}
    try:
        from .mailbox import list_mail_messages, sync_gmail_metadata
    except Exception:
        return [], quotes, raw_texts

    # Keep the dashboard button fast. Refresh a small recent slice from Gmail,
    # then mine the local mailbox cache for the wider 180-day history that was
    # already synced by background jobs or prior runs.
    sync_limit = _env_int("SHIMS_INVENTORY_GMAIL_SYNC_LIMIT", 5, lo=0, hi=max(0, min(max_mail, 100)))
    if sync_limit:
        try:
            sync_gmail_metadata(query="newer_than:30d", max_results=sync_limit)
            sync_gmail_metadata(query="in:sent newer_than:30d", max_results=max(5, sync_limit // 2))
        except Exception:
            pass
    try:
        cached = list_mail_messages(limit=max(2000, max_mail), provider="gmail")
    except Exception:
        cached = []

    # A newest-N cut never reaches vendor rate history in a high-volume
    # mailbox (~150 mails/week here). Keep the recent slice, plus ANY mail
    # whose subject/snippet is commercial (payments, POs, quotations…) no
    # matter how old — that is where vendor rates live.
    _COMMERCIAL_RE = re.compile(
        r"payment|quotation|invoice|proforma|purchase\s?order|\bpo[- ]?no|"
        r"\brate|price|quote|ledger|dispatch|challan|outstanding|overdue|"
        r"balance|supply|supplied|tender|enquiry|inquiry", re.I)
    cutoff = (datetime.now() - timedelta(days=14)).strftime("%Y-%m-%d")
    cached = [m for m in cached
              if str(m.get("received_at") or "")[:10] >= cutoff
              or _COMMERCIAL_RE.search(f"{m.get('subject') or ''} {m.get('snippet') or ''}")][: max(2000, max_mail)]

    inbox_messages: list[dict[str, Any]] = []
    sent_messages: list[dict[str, Any]] = []
    for m in cached:
        labels = {str(x).upper() for x in (m.get("labels") or [])}
        if "SENT" in labels:
            sent_messages.append(m)
        else:
            inbox_messages.append(m)

    for m in inbox_messages:
        vendor = (m.get("sender") or "unknown").strip()
        ref = m.get("thread_id") or m.get("external_id") or m.get("id") or ""
        url = f"https://mail.google.com/mail/u/1/#all/{ref}" if ref else ""
        date = str(m.get("received_at") or "")[:10]
        entry = vendors.setdefault(vendor, {
            "vendor": vendor, "channels": {"gmail"}, "last_at": "",
            "products": [], "equipment": [], "services": [],
            "message_count": 0, "url": url})
        entry["message_count"] += 1
        if date > entry["last_at"]:
            entry["last_at"] = date
        text = f"{m.get('subject') or ''}\n{m.get('snippet') or ''}"
        raw_texts.setdefault(vendor, []).append(text[:400])
        for row in _extract_items(text, source="gmail", url=url, date=date):
            key = row["category"]
            if row["item"] not in [r["item"] for r in entry[key]]:
                entry[key].append(row)
    for m in sent_messages:
        # Outbound quotes: the counterparty is the recipient.
        counter = (m.get("recipients") or m.get("sender") or "unknown").strip()
        ref = m.get("thread_id") or m.get("external_id") or m.get("id") or ""
        url = f"https://mail.google.com/mail/u/1/#all/{ref}" if ref else ""
        date = str(m.get("received_at") or "")[:10]
        text = f"{m.get('subject') or ''}\n{m.get('snippet') or ''}"
        raw_texts.setdefault(f"You → {counter}", []).append(text[:400])
        for row in _extract_items(text, source="gmail", url=url, date=date):
            if row["rate"]:
                row["counterparty"] = counter
                quotes.append(row)
    return list(vendors.values()), quotes, raw_texts


def _llm_enrich(raw_texts: dict[str, list[str]]) -> dict[str, dict[str, list[dict[str, Any]]]]:
    """LLM pass: per vendor, extract clean {item, rate, qty} rows from their
    messages. Only vendors with offer/price context, capped — the heuristic
    stays the source of truth when the engine is busy or the parse fails."""
    import os
    if (os.getenv("SHIMS_INVENTORY_LLM") or "on").strip().lower() in {"off", "0", "false", "no"}:
        return {}
    try:
        from .native_engine import get_engine
        engine = get_engine()
        if not engine.loaded_model_id():
            return {}
    except Exception:
        return {}

    def _interesting(texts: list[str]) -> bool:
        blob = " ".join(texts)
        return bool(_OFFER_RE.search(blob) or _PRICE_RE.search(blob))

    candidates = [v for v, t in raw_texts.items() if _interesting(t)][:6]
    out: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for vendor in candidates:
        texts = [t for t in raw_texts[vendor] if t.strip()][:8]
        if not texts:
            continue
        listing = "\n---\n".join(texts)
        prompt = (
            "From these messages by one vendor, extract every offered product/raw material, "
            "equipment item, and service, with the quoted rate (if any) and quantity (if any). "
            "Reply with ONLY a JSON object like "
            '{"products":[{"item":"...","rate":"...","qty":"..."}],"equipment":[...],"services":[...]}. '
            "Use empty strings for missing rate/qty. No prose, no markdown.\n\n" + listing)
        try:
            result = engine.chat_raw([{"role": "user", "content": prompt}],
                                     max_tokens=1000, timeout=300.0)
        except Exception:
            continue
        text = result.get("content") or ""
        start, end = text.find("{"), text.rfind("}")
        if start < 0 or end <= start:
            continue
        try:
            data = json.loads(text[start:end + 1])
        except Exception:
            continue
        cats: dict[str, list[dict[str, Any]]] = {}
        for key in ("products", "equipment", "services"):
            rows = []
            for item in (data.get(key) or [])[:10]:
                if not isinstance(item, dict) or not item.get("item"):
                    continue
                rows.append({
                    "category": key,
                    "item": str(item.get("item"))[:160],
                    "rate": str(item.get("rate") or "")[:120],
                    "qty": str(item.get("qty") or "")[:40],
                    "date": "", "source": "llm", "url": "", "offer": bool(item.get("rate")),
                })
            cats[key] = rows
        out[vendor] = cats
    return out


def build_inventory(*, wa_limit: int = 500, mail_limit: int = 150) -> dict[str, Any]:
    """Vendor-wise products / equipment / services index with rates."""
    wa_vendors, wa_quotes, wa_texts = _gather_whatsapp(wa_limit)
    gm_vendors, gm_quotes, gm_texts = _gather_gmail(mail_limit)
    raw_texts = {**wa_texts}
    for vendor, texts in gm_texts.items():
        raw_texts.setdefault(vendor, []).extend(texts)
    llm_rows = _llm_enrich(raw_texts)

    vendors: dict[str, dict[str, Any]] = {}
    for entry in wa_vendors + gm_vendors:
        cur = vendors.setdefault(entry["vendor"], entry)
        if cur is entry:
            continue
        cur["channels"] = cur["channels"] | entry["channels"]
        cur["message_count"] += entry["message_count"]
        if entry["last_at"] > cur["last_at"]:
            cur["last_at"] = entry["last_at"]
        if entry["url"] and not cur["url"]:
            cur["url"] = entry["url"]
        for key in ("products", "equipment", "services"):
            seen = {r["item"] for r in cur[key]}
            cur[key] += [r for r in entry[key] if r["item"] not in seen]
    # LLM enrichment: append clean rows the heuristic missed (dedupe by item).
    for vendor, cats in llm_rows.items():
        cur = vendors.get(vendor)
        if not cur:
            continue
        for key in ("products", "equipment", "services"):
            seen = {r["item"].lower()[:60] for r in cur[key]}
            cur[key] += [r for r in cats.get(key, [])
                         if r["item"].lower()[:60] not in seen]

    vendor_list = []
    for v in vendors.values():
        v["channels"] = sorted(v["channels"])
        for key in ("products", "equipment", "services"):
            v[key] = v[key][:12]
        v["indexed"] = sum(len(v[k]) for k in ("products", "equipment", "services"))
        v["rated"] = sum(1 for k in ("products", "equipment", "services")
                         for r in v[k] if r["rate"])
        vendor_list.append(v)
    vendor_list.sort(key=lambda v: (v["rated"], v["indexed"], v["message_count"]), reverse=True)

    # My quoted rates: what WE offered each counterparty (deduped, capped).
    quotes = wa_quotes + gm_quotes
    seen_q: set[str] = set()
    my_quotes: list[dict[str, Any]] = []
    for q in sorted(quotes, key=lambda r: r.get("date", ""), reverse=True):
        key = (q.get("counterparty", "") + q["item"] + q["rate"]).lower()[:120]
        if key in seen_q:
            continue
        seen_q.add(key)
        my_quotes.append(q)
    my_quotes = my_quotes[:30]

    board = {
        "generatedAt": _now_iso(),
        "vendor_count": len(vendor_list),
        "vendors": vendor_list,
        "my_quotes": my_quotes,
        "llm_enriched": sorted(llm_rows.keys()),
    }
    INVENTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
    INVENTORY_PATH.write_text(json.dumps(board, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"ok": True, "vendor_count": len(vendor_list),
            "rated_rows": sum(v["rated"] for v in vendor_list),
            "my_quotes": len(my_quotes), "llm_enriched": len(llm_rows),
            "path": str(INVENTORY_PATH)}


_FINALIZED_RE = re.compile(
    r"\b(final\s?\d+\s?%|finalized|finalised|confirmed\s?(rate|price)|"
    r"as\s?agreed|purchase\s?order|po[- ]?no|po[- ]?\d|invoice|tax\s?invoice|"
    r"ledger|payment\s?(received|pending|released)|balance\s?payment|"
    r"short\s?payment|overdue)\b", re.I)


def latest_inventory() -> dict[str, Any]:
    try:
        if INVENTORY_PATH.is_file():
            return json.loads(INVENTORY_PATH.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {"generatedAt": "", "vendor_count": 0, "vendors": [], "my_quotes": []}


def export_xlsx(out_path: str | Path | None = None) -> dict[str, Any]:
    """Write the vendor index to a 3-sheet Excel workbook (openpyxl)."""
    import openpyxl

    board = latest_inventory()
    if not board.get("vendors"):
        build_inventory()
        board = latest_inventory()

    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = Path(out_path) if out_path else (
        EXPORT_DIR / f"vendor_inventory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")

    wb = openpyxl.Workbook()
    headers = ["Vendor", "Item / Offer", "Rate", "Qty", "Date", "Source", "Channels", "Link"]

    summary = wb.active
    summary.title = "Vendors"
    summary.append(["Vendor", "Products", "Equipment", "Services", "Rated rows",
                    "Messages", "Last activity", "Channels", "Link"])
    for v in board["vendors"]:
        summary.append([v["vendor"], len(v["products"]), len(v["equipment"]),
                        len(v["services"]), v["rated"], v["message_count"],
                        v["last_at"], ", ".join(v["channels"]), v["url"]])

    for key, title in (("products", "Products"), ("equipment", "Equipment"),
                       ("services", "Services")):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for v in board["vendors"]:
            for row in v[key]:
                ws.append([v["vendor"], row["item"], row["rate"], row["qty"],
                           row["date"], row["source"], ", ".join(v["channels"]), row["url"]])
        for col, width in zip("ABCDEFGH", (28, 60, 18, 10, 12, 10, 14, 46)):
            ws.column_dimensions[col].width = width

    quotes_ws = wb.create_sheet("My Quotes")
    quotes_ws.append(["Counterparty", "Item / Offer", "Rate I quoted", "Qty",
                      "Date", "Source", "Link"])
    for q in board.get("my_quotes") or []:
        quotes_ws.append([q.get("counterparty", ""), q["item"], q["rate"], q["qty"],
                          q["date"], q["source"], q["url"]])
    for col, width in zip("ABCDEFG", (30, 60, 18, 10, 12, 10, 46)):
        quotes_ws.column_dimensions[col].width = width
    for col, width in zip("ABCDEFGHI", (30, 10, 10, 10, 10, 10, 14, 14, 46)):
        summary.column_dimensions[col].width = width

    wb.save(path)
    return {"ok": True, "path": str(path), "vendors": board["vendor_count"],
            "generatedAt": board.get("generatedAt", "")}
