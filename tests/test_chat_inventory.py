"""Chat inventory: extraction, vendor merge, xlsx export."""
from __future__ import annotations

import json

from shared import chat_inventory
from shared.chat_inventory import _PRICE_RE, _extract_items, build_inventory, export_xlsx, latest_inventory


def test_price_patterns():
    assert "Rs 450/kg" in _PRICE_RE.findall("available at Rs 450/kg ready stock")[0]
    assert _PRICE_RE.findall("price ₹2.5 lakh, immediate dispatch")[0] == "₹2.5 lakh"
    assert _PRICE_RE.findall("transport @ 950/- per ton")[0] == "@ 950/-"
    found = _PRICE_RE.findall("HPLC column USD 320, blender 1.2 crore")
    assert any("320" in f for f in found) and any("crore" in f for f in found)


def test_extract_categories_and_rates():
    rows = _extract_items(
        "READY STOCK: 8-Hydroxyquinoline powder at Rs 450/kg. "
        "Also supplying blender machine ₹2.5 lakh. We do third-party testing service.",
        source="whatsapp", url="https://wa.me/919999999999", date="2026-08-04")
    cats = {r["category"] for r in rows}
    assert "products" in cats and "equipment" in cats and "services" in cats
    prod = next(r for r in rows if r["category"] == "products")
    assert "450" in prod["rate"]
    assert prod["url"].startswith("https://wa.me/")


def test_build_and_export(tmp_path, monkeypatch):
    monkeypatch.setattr(chat_inventory, "INVENTORY_PATH", tmp_path / "inv.json")
    monkeypatch.setattr(chat_inventory, "EXPORT_DIR", tmp_path / "exports")
    monkeypatch.setattr(chat_inventory, "_llm_enrich", lambda texts: {})
    monkeypatch.setattr(chat_inventory, "_gather_whatsapp", lambda limit: ([{
        "vendor": "Chem Supplier", "channels": {"whatsapp"}, "last_at": "2026-08-04",
        "products": [{"category": "products", "item": "quinoline powder", "rate": "Rs 450/kg",
                      "qty": "25 kg", "date": "2026-08-04", "source": "whatsapp",
                      "url": "https://wa.me/9111", "offer": True}],
        "equipment": [], "services": [], "message_count": 3, "url": "https://wa.me/9111"}],
        [{"category": "products", "item": "quoted resin", "rate": "Rs 300/kg", "qty": "",
          "date": "2026-08-04", "source": "whatsapp", "url": "https://wa.me/9222",
          "offer": True, "counterparty": "91999"}],
        {}))
    monkeypatch.setattr(chat_inventory, "_gather_gmail", lambda max_mail: ([], [], {}))
    result = build_inventory()
    assert result["ok"] and result["vendor_count"] == 1 and result["rated_rows"] == 1
    assert result["my_quotes"] == 1
    board = latest_inventory()
    v = board["vendors"][0]
    assert v["vendor"] == "Chem Supplier"
    assert v["channels"] == ["whatsapp"]
    assert board["my_quotes"][0]["counterparty"] == "91999"
    out = export_xlsx()
    assert out["ok"] and out["path"].endswith(".xlsx")
    import openpyxl
    wb = openpyxl.load_workbook(out["path"])
    assert set(wb.sheetnames) == {"Vendors", "Products", "Equipment", "Services", "My Quotes"}
    assert wb["Products"].max_row == 2  # header + 1 row
    assert "450" in str(wb["Products"].cell(row=2, column=3).value)
    assert wb["My Quotes"].max_row == 2
    assert "300" in str(wb["My Quotes"].cell(row=2, column=3).value)
